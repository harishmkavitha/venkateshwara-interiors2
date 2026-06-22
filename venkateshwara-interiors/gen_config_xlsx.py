#!/usr/bin/env python3
# Builds config.xlsx — the editable, reusable client configuration.
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(ROOT, "config.xlsx")

TEAL = "0E2A2B"; BRASS = "C9A24B"; IVORY = "F6F1E7"; LIGHT = "EAE2D2"
thin = Side(style="thin", color="CBC3B2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------- Settings sheet ----------------
ws = wb.active
ws.title = "Settings"
ws.sheet_view.showGridLines = False

title_font = Font(name="Arial", size=15, bold=True, color=IVORY)
hdr_font   = Font(name="Arial", size=11, bold=True, color=IVORY)
key_font   = Font(name="Arial", size=10, bold=True, color=TEAL)
val_font   = Font(name="Arial", size=10, color="222222")
note_font  = Font(name="Arial", size=9, italic=True, color="6A6A6A")
teal_fill  = PatternFill("solid", fgColor=TEAL)
brass_fill = PatternFill("solid", fgColor=BRASS)
light_fill = PatternFill("solid", fgColor=LIGHT)

ws.merge_cells("A1:C1")
ws["A1"] = "VENKATESHWARA INTERIORS  —  Website Settings"
ws["A1"].font = title_font; ws["A1"].fill = teal_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:C2")
ws["A2"] = "Edit the Value column only. Keep the Key column exactly as-is. Re-host (or re-export to config.js) to update the live site."
ws["A2"].font = note_font; ws["A2"].fill = light_fill
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 22

# header row
for col, label in zip("ABC", ["Key (do not change)", "Value (edit me)", "Notes"]):
    c = ws[f"{col}3"]; c.value = label; c.font = hdr_font; c.fill = brass_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1); c.border = border
ws.row_dimensions[3].height = 22

settings = [
 ("businessName","Venkateshwara Interiors","Brand name shown in header, footer, titles"),
 ("tagline","Crafting Soulful Indian Homes","Short brand tagline"),
 ("shortAbout","A premium interior design studio shaping modern Indian homes — from modular kitchens and pooja rooms to complete turnkey interiors, rooted in craft and built to last.","Footer 'about' paragraph"),
 ("establishedYear","2009","Year founded (badges / about)"),
 ("phone","+91 98840 12345","Display phone number"),
 ("phoneRaw","919884012345","Digits only — used for tel: and click-to-call"),
 ("whatsapp","919884012345","Digits only — used for WhatsApp link"),
 ("email","hello@venkateshwarainteriors.in","Contact email"),
 ("websiteUrl","https://www.venkateshwarainteriors.in","Full website URL"),
 ("addressLine1","No. 24, Sannathi Street, Velachery","Address line 1"),
 ("addressLine2","Chennai, Tamil Nadu 600042, India","Address line 2"),
 ("mapEmbed","https://www.google.com/maps?q=Velachery,Chennai&output=embed","Google Maps EMBED url (used in the contact iframe)"),
 ("mapLink","https://maps.google.com/?q=Velachery,Chennai","Google Maps link (used by the Map action button)"),
 ("gstin","33ABCDE1234F1Z5","GST registration number"),
 ("cin","U74999TN2009PTC012345","Company CIN (optional)"),
 ("hours","Mon – Sat: 10:00 AM – 8:00 PM","Working hours line"),
 ("hoursNote","Sunday by appointment","Secondary hours note"),
 ("facebook","https://facebook.com/","Facebook page URL (blank to hide)"),
 ("instagram","https://instagram.com/","Instagram URL (blank to hide)"),
 ("youtube","https://youtube.com/","YouTube URL (blank to hide)"),
 ("linkedin","https://linkedin.com/","LinkedIn URL (blank to hide)"),
 ("pinterest","https://pinterest.com/","Pinterest URL (blank to hide)"),
 ("primarySocial","instagram","Which social icon shows in the mobile action bar"),
 ("copyrightName","Venkateshwara Interiors","Name in the footer copyright line"),
]
r = 4
for key, val, note in settings:
    ws[f"A{r}"] = key; ws[f"A{r}"].font = key_font
    ws[f"B{r}"] = val; ws[f"B{r}"].font = val_font
    ws[f"C{r}"] = note; ws[f"C{r}"].font = note_font
    fill = light_fill if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col in "ABC":
        cell = ws[f"{col}{r}"]; cell.fill = fill; cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=(col=="C"))
    r += 1

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 66
ws.column_dimensions["C"].width = 50
ws.freeze_panes = "A4"

# ---------------- Guidelines sheet ----------------
gw = wb.create_sheet("Guidelines")
gw.sheet_view.showGridLines = False
gw.merge_cells("A1:D1")
gw["A1"] = "Website Best-Practice Sections (reference)"
gw["A1"].font = title_font; gw["A1"].fill = teal_fill
gw["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
gw.row_dimensions[1].height = 30
for col, label in zip("ABCD", ["No.","Website Item","Meaning / Purpose","Common Contents"]):
    c = gw[f"{col}2"]; c.value = label; c.font = hdr_font; c.fill = brass_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1); c.border = border
gw.row_dimensions[2].height = 20

GUIDE = [
 (1,"Top Bar","Small strip above the main header","Phone, email, address, social icons"),
 (2,"Header","Main top area of website","Logo, menu, search, button"),
 (3,"Logo Area","Shows brand identity","Logo image, company name"),
 (4,"Navigation Bar / Menu","Helps users move between pages","Home, About, Services, Contact"),
 (5,"Dropdown Menu","Sub-menu under main item","Service list, page list"),
 (6,"Mega Menu","Large dropdown menu","Multiple columns, links"),
 (10,"CTA Button in Header","Main action button","Book Now, Get Quote, Contact"),
 (11,"Mobile Menu","Menu for small screens","Hamburger icon, slide menu"),
 (12,"Hero Section","First big section after header","Heading, text, image, button"),
 (13,"Banner Section","Visual introduction area","Background image, title, subtitle"),
 (15,"Breadcrumb","Shows page location","Home > Services > Details"),
 (22,"Grid Layout","Arranges content in boxes","Service cards, gallery items"),
 (23,"About Section","Explains the business","Intro, mission, vision, image"),
 (24,"Services Section","Shows offered services","Service cards, icons, descriptions"),
 (25,"Service Details","Explains one service deeply","Features, benefits, process"),
 (26,"Features Section","Shows key advantages","Quality, team, warranty"),
 (27,"Why Choose Us","Builds trust","Experience, benefits, statistics"),
 (28,"Process Section","Shows working steps","Step 1, Step 2, Step 3"),
 (33,"Portfolio / Gallery","Shows previous work","Images, categories, lightbox"),
 (34,"Gallery Section","Displays photos / videos","Image grid, lightbox"),
 (38,"Testimonials","Shows customer reviews","Review text, name, rating"),
 (41,"FAQ Section","Answers common questions","Question, answer, accordion"),
 (46,"Call-to-Action","Pushes user to act","Heading, short text, button"),
 (49,"Contact Section","Shows contact options","Address, phone, email, form"),
 (50,"Contact Form","Collects user messages","Name, email, subject, message"),
 (52,"Map Section","Shows physical location","Google Map, address"),
 (53,"Statistics","Shows achievements","Clients, projects, years"),
 (60,"Accordion","Expand / collapse content","FAQ, service info"),
 (63,"Lightbox","Opens images / videos larger","Gallery image preview"),
 (68,"Social Media Icons","Links to social profiles","Facebook, Instagram, LinkedIn"),
 (69,"Footer","Bottom section of website","Logo, links, contact, newsletter"),
 (75,"Copyright Bar","Very bottom line","© Year Company. All Rights Reserved"),
 (76,"Privacy Policy","Legal / privacy information","Data usage, cookies, rights"),
 (77,"Terms & Conditions","Website rules","Terms, limitations, policies"),
 (78,"404 Page","Error page for missing links","Error message, back home"),
 (79,"Sitemap Page","Lists website pages","All page links"),
 (80,"Back-to-Top Button","Scrolls page to top","Arrow button"),
 (81,"Sticky Header","Header fixed while scrolling","Logo, menu, CTA"),
 (82,"Preloader","Loading animation","Spinner, logo animation"),
 (83,"Animation Effects","Adds movement","Fade, slide, zoom"),
 (84,"Responsive Layout","Adapts to screen size","Desktop, tablet, mobile"),
 (89,"SEO Meta Tags","Helps search engines","Title, description, keywords"),
 (90,"Favicon","Browser tab icon","Brand icon"),
]
r = 3
for no, item, mean, cont in GUIDE:
    gw[f"A{r}"] = no; gw[f"B{r}"] = item; gw[f"C{r}"] = mean; gw[f"D{r}"] = cont
    fill = light_fill if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col, fnt in zip("ABCD", [val_font, key_font, val_font, note_font]):
        cell = gw[f"{col}{r}"]; cell.font = fnt; cell.fill = fill; cell.border = border
        cell.alignment = Alignment(horizontal=("center" if col=="A" else "left"), vertical="center",
                                   indent=(0 if col=="A" else 1), wrap_text=True)
    r += 1
gw.column_dimensions["A"].width = 7
gw.column_dimensions["B"].width = 24
gw.column_dimensions["C"].width = 38
gw.column_dimensions["D"].width = 40
gw.freeze_panes = "A3"

wb.save(OUT)
print("saved", OUT)
